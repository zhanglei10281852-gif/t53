from datetime import date
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import (
    InventoryItemWithExpiry,
    ProductSummary,
    ExpiryStatus,
)
from calculator import get_alert_items, summarize_by_product, get_totals


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
EXPIRED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
RED_ALERT_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row_num: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_row(ws, row_num: int, num_cols: int, fill=None):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fill:
            cell.fill = fill


def auto_width(ws, num_cols: int, min_width=10, max_width=30):
    for col in range(1, num_cols + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            cell_value = str(row[0]) if row[0] is not None else ""
            if len(cell_value) > max_len:
                max_len = len(cell_value)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, max_width)


def generate_inventory_sheet(
    wb: Workbook, items: List[InventoryItemWithExpiry]
):
    ws = wb.active
    ws.title = "库存明细"

    headers = [
        "品名",
        "规格",
        "批次号",
        "生产日期",
        "保质期月数",
        "库存数量",
        "存放位置",
        "单价(元)",
        "到期日期",
        "剩余天数",
        "效期状态",
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header(ws, 1, len(headers))

    row_num = 2
    for item in items:
        ws.cell(row=row_num, column=1, value=item.name)
        ws.cell(row=row_num, column=2, value=item.specification)
        ws.cell(row=row_num, column=3, value=item.batch_number)
        ws.cell(row=row_num, column=4, value=item.production_date.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=5, value=item.shelf_life_months)
        ws.cell(row=row_num, column=6, value=item.quantity)
        ws.cell(row=row_num, column=7, value=item.location)
        ws.cell(row=row_num, column=8, value=item.unit_price)
        ws.cell(row=row_num, column=9, value=item.expiry_date.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=10, value=item.days_remaining)
        ws.cell(row=row_num, column=11, value=item.expiry_status.value)

        fill = None
        if item.expiry_status == ExpiryStatus.EXPIRED:
            fill = EXPIRED_FILL
        elif item.expiry_status == ExpiryStatus.RED_ALERT:
            fill = RED_ALERT_FILL
        style_data_row(ws, row_num, len(headers), fill)

        row_num += 1

    product_count, batch_count, total_value = get_totals(items)

    total_row = row_num
    ws.cell(row=total_row, column=1, value="汇总")
    ws.cell(row=total_row, column=2, value=f"品种总数: {product_count}")
    ws.cell(row=total_row, column=3, value=f"批次总数: {batch_count}")
    ws.cell(row=total_row, column=4, value=f"总金额: {total_value:.2f}元")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=1)
    ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=3)
    ws.merge_cells(start_row=total_row, start_column=4, end_row=total_row, end_column=11)

    auto_width(ws, len(headers))
    ws.row_dimensions[1].height = 25


def generate_alert_sheet(
    wb: Workbook, items: List[InventoryItemWithExpiry], report_date: date
):
    ws = wb.create_sheet("效期预警清单")

    title = f"效期预警清单  生成日期：{report_date.strftime('%Y年%m月%d日')}"
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "品名",
        "规格",
        "批次号",
        "生产日期",
        "到期日期",
        "剩余天数",
        "效期状态",
        "库存数量",
        "存放位置",
    ]

    alert_items = get_alert_items(items)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 35

    for col, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=header)
    style_header(ws, 3, len(headers))

    row_num = 4
    for item in alert_items:
        ws.cell(row=row_num, column=1, value=item.name)
        ws.cell(row=row_num, column=2, value=item.specification)
        ws.cell(row=row_num, column=3, value=item.batch_number)
        ws.cell(row=row_num, column=4, value=item.production_date.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=5, value=item.expiry_date.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=6, value=item.days_remaining)
        ws.cell(row=row_num, column=7, value=item.expiry_status.value)
        ws.cell(row=row_num, column=8, value=item.quantity)
        ws.cell(row=row_num, column=9, value=item.location)

        fill = None
        if item.expiry_status == ExpiryStatus.EXPIRED:
            fill = EXPIRED_FILL
        elif item.expiry_status == ExpiryStatus.RED_ALERT:
            fill = RED_ALERT_FILL
        style_data_row(ws, row_num, len(headers), fill)

        row_num += 1

    auto_width(ws, len(headers))
    ws.row_dimensions[3].height = 25


def generate_summary_sheet(
    wb: Workbook, items: List[InventoryItemWithExpiry]
):
    ws = wb.create_sheet("品名汇总统计")

    summaries = summarize_by_product(items)

    headers = [
        "品名",
        "总库存",
        "总金额(元)",
        "批次数",
        "最早到期日期",
        "最早批次效期状态",
        "库存状态",
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header(ws, 1, len(headers))

    row_num = 2
    for summary in summaries:
        ws.cell(row=row_num, column=1, value=summary.name)
        ws.cell(row=row_num, column=2, value=summary.total_quantity)
        ws.cell(row=row_num, column=3, value=round(summary.total_value, 2))
        ws.cell(row=row_num, column=4, value=summary.batch_count)
        ws.cell(row=row_num, column=5, value=summary.earliest_expiry_date.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=6, value=summary.earliest_expiry_status.value)
        ws.cell(row=row_num, column=7, value="缺货" if summary.is_out_of_stock else "正常")

        fill = None
        if summary.earliest_expiry_status == ExpiryStatus.EXPIRED:
            fill = EXPIRED_FILL
        elif summary.earliest_expiry_status == ExpiryStatus.RED_ALERT:
            fill = RED_ALERT_FILL
        style_data_row(ws, row_num, len(headers), fill)

        row_num += 1

    auto_width(ws, len(headers))
    ws.row_dimensions[1].height = 25


def generate_report(
    items: List[InventoryItemWithExpiry],
    output_path: str,
    report_date: date,
):
    wb = Workbook()
    generate_inventory_sheet(wb, items)
    generate_alert_sheet(wb, items, report_date)
    generate_summary_sheet(wb, items)
    wb.save(output_path)


def print_alert_terminal(items: List[InventoryItemWithExpiry]):
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel

    console = Console()
    alert_items = get_alert_items(items)

    expired_count = sum(
        1 for item in alert_items if item.expiry_status == ExpiryStatus.EXPIRED
    )
    red_alert_count = sum(
        1 for item in alert_items if item.expiry_status == ExpiryStatus.RED_ALERT
    )

    console.print(
        Panel.fit(
            f"[bold red]已过期: {expired_count} 条[/bold red]  |  "
            f"[bold orange_red1]近效期预警(≤90天): {red_alert_count} 条[/bold orange_red1]",
            title="[bold yellow]效期预警清单[/bold yellow]",
            border_style="red",
        )
    )

    if not alert_items:
        console.print("[green]没有需要预警的药品[/green]")
        return

    table = Table(show_header=True, header_style="bold blue")
    table.add_column("序号", justify="center", style="dim", width=6)
    table.add_column("品名", style="bold")
    table.add_column("规格")
    table.add_column("批次号")
    table.add_column("到期日期", justify="center")
    table.add_column("剩余天数", justify="right")
    table.add_column("效期状态", justify="center")
    table.add_column("库存数量", justify="right")
    table.add_column("存放位置")

    for idx, item in enumerate(alert_items, start=1):
        if item.expiry_status == ExpiryStatus.EXPIRED:
            status_str = f"[bold red]{item.expiry_status.value}[/bold red]"
            days_str = f"[bold red]{item.days_remaining}[/bold red]"
        else:
            status_str = f"[bold orange_red1]{item.expiry_status.value}[/bold orange_red1]"
            days_str = f"[bold orange_red1]{item.days_remaining}[/bold orange_red1]"

        table.add_row(
            str(idx),
            item.name,
            item.specification,
            item.batch_number,
            item.expiry_date.strftime("%Y-%m-%d"),
            days_str,
            status_str,
            str(item.quantity),
            item.location,
        )

    console.print(table)
